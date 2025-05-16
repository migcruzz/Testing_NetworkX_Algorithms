import re
from multiprocessing import Process, Manager, freeze_support
from pathlib import Path

import networkx as nx
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from Algorithms.bi_astar import bidirectional_astar
from Algorithms.ida_star import idastar_path
from Algorithms.rtaa_star import rtaa_star_path
from Algorithms.sma_star import sma_star_path
from CsvProcessor.generator import generate_graph_from_csv
from CsvProcessor.processor import CSVLinearCombiner
from Graphs.graphs import GraphPlotter
from Testers.bi_astar import AStarVsBidirectionalComparison
from Testers.d_star_lite import DStarLiteVsAStarComparison
from Testers.ida_star import IDAStarVsAStarComparison
from Testers.rtaa_star import RTAAStarVsAStarComparison
from Testers.sma_star import SMAStarVsAStarComparison
from config import MEMORY_LIMIT, LOOKAHEAD, MOVELIMIT, N_MODIFICATIONS, SOURCE, TARGET, EXCEL_FILE, DIRECTED, \
    CSV_PROCESSED_PATH, CSV_PATH


# ─── Excel helpers ────────────────────────────────────────────────────────────
def sanitize_sheet_name(name: str) -> str:
    return re.sub(r'[:\\/?*\[\]]', "", name)[:31]


def append_sheet(result_dict, sheet_name, excel_path=EXCEL_FILE):
    df = pd.DataFrame(result_dict.items(), columns=["Metric", "Value"])
    sheet_name = sanitize_sheet_name(sheet_name)
    mode = "a" if Path(excel_path).exists() else "w"

    with pd.ExcelWriter(
            excel_path,
            engine="openpyxl",
            mode=mode,
            if_sheet_exists=None if mode == "w" else "replace",
    ) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        bold = Font(bold=True)
        centre = Alignment(horizontal="center")

        for cell in ws[1]:
            cell.font = bold
            cell.alignment = centre

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = centre

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(c.value)) if c.value is not None else 0
                for c in ws[col_letter]
            )
            ws.column_dimensions[col_letter].width = max_len + 2


# Worker wrappers

def run_astar(graph, n_mods_unused, shared):
    try:
        path = nx.astar_path(graph, SOURCE, TARGET, weight="weight")
        cost = sum(graph[u][v]["weight"] for u, v in zip(path, path[1:]))

        shared["A*"] = {"Cost": cost, "Path length": len(path) - 1}

    except Exception as exc:
        print("Erro no worker A* :", exc)


def run_bidirectional(graph, n_mods, shared):
    tester = AStarVsBidirectionalComparison(graph, SOURCE, TARGET, n_mods)
    shared["Bidirectional A*"] = tester.run_all()


def run_dstar(di_graph, n_mods, shared):
    tester = DStarLiteVsAStarComparison(di_graph, SOURCE, TARGET, n_mods)
    shared["D* Lite"] = tester.run_all()


def run_idastar(graph, n_mods, shared):
    tester = IDAStarVsAStarComparison(graph, SOURCE, TARGET, n_mods)
    shared["IDA*"] = tester.run_all()


def run_rtaa(graph, lookahead, move_limit, n_mods, shared):
    key = f"RTAA* (L={lookahead}, M={move_limit})"
    tester = RTAAStarVsAStarComparison(graph, SOURCE, TARGET, lookahead, move_limit, n_mods)
    shared[key] = tester.run_all()


def run_sma(graph, memory_limit, n_mods, shared):
    tester = SMAStarVsAStarComparison(graph, SOURCE, TARGET, memory_limit=memory_limit, n_modifications=n_mods)
    shared["SMA*"] = tester.run_all()


def launch(worker, *args):
    worker(*args)


# Main
if __name__ == "__main__":
    freeze_support()

    combiner = CSVLinearCombiner(
        file_path=Path(CSV_PATH),
        weights=[0.33, 0.33, 0.33],  # w0, w1, w2
        encoding="utf-8",
    )

    df_out = combiner.process()
    df_out.to_csv(Path(CSV_PROCESSED_PATH),
                  index=False,
                  encoding="utf-8")

    base_graph = generate_graph_from_csv(CSV_PROCESSED_PATH, "combined_cost", DIRECTED)
    if not nx.has_path(base_graph, SOURCE, TARGET):
        print("No path between source and target. Aborting.")
        exit(1)

    directed_graph = base_graph.to_directed()
    base_plotter = GraphPlotter(base_graph)
    directed_plotter = GraphPlotter(directed_graph)

    manager = Manager()
    shared_results = manager.dict()

    try:
        astar_path = nx.astar_path(base_graph, SOURCE, TARGET, weight="weight")
        cost = sum(base_graph[u][v]["weight"] for u, v in zip(astar_path, astar_path[1:]))

        base_plotter.draw(
            source=SOURCE,
            target=TARGET,
            path=astar_path,
            metrics={"Cost": cost},
            output_path="Graphs/Plots/path_astar.png"
        )
    except Exception as e:
        print("Erro ao desenhar caminho A*:", e)

    processes = [
        Process(target=launch, args=(run_astar, base_graph, N_MODIFICATIONS, shared_results)),
        Process(target=launch, args=(run_bidirectional, base_graph, N_MODIFICATIONS, shared_results)),
        Process(target=launch, args=(run_dstar, directed_graph, N_MODIFICATIONS, shared_results)),
        Process(target=launch, args=(run_idastar, base_graph, N_MODIFICATIONS, shared_results)),
        Process(target=launch, args=(run_rtaa, base_graph, LOOKAHEAD, MOVELIMIT, N_MODIFICATIONS, shared_results)),
        Process(target=launch, args=(run_sma, base_graph, MEMORY_LIMIT, N_MODIFICATIONS, shared_results)),
    ]

    for p in processes:
        p.start()
    for p in processes:
        p.join()

    # Draw graphs with metrics (post-process)
    try:
        _, bidir_path, _ = bidirectional_astar(base_graph, SOURCE, TARGET)
        base_plotter.draw(SOURCE, TARGET, path=bidir_path, metrics=shared_results["Bidirectional A*"],
                          output_path="Graphs/Plots/path_bidirectional_astar.png")
    except:
        pass

    try:
        dstar_path = nx.astar_path(directed_graph, SOURCE, TARGET, weight="weight")
        directed_plotter.draw(SOURCE, TARGET, path=dstar_path, metrics=shared_results["D* Lite"],
                              output_path="Graphs/Plots/path_dstar_lite.png")
    except:
        pass

    try:
        ida_path = idastar_path(base_graph, SOURCE, TARGET)
        base_plotter.draw(SOURCE, TARGET, path=ida_path, metrics=shared_results["IDA*"],
                          output_path="Graphs/Plots/path_idastar.png")
    except:
        pass

    try:
        rtaa_path = rtaa_star_path(base_graph, SOURCE, TARGET, lookahead=LOOKAHEAD, move_limit=MOVELIMIT)
        base_plotter.draw(SOURCE, TARGET, path=rtaa_path,
                          metrics=shared_results[f"RTAA* (L={LOOKAHEAD}, M={MOVELIMIT})"],
                          output_path="Graphs/Plots/path_rtaa_star.png")
    except:
        pass

    try:
        sma_path = sma_star_path(base_graph, SOURCE, TARGET, memory_limit=MEMORY_LIMIT)
        base_plotter.draw(SOURCE, TARGET, path=sma_path, metrics=shared_results["SMA*"],
                          output_path="Graphs/Plots/path_sma_star.png")
    except:
        pass

    for algorithm, metrics in shared_results.items():
        append_sheet(metrics, algorithm, EXCEL_FILE)

    print(f"All metrics written to {EXCEL_FILE}")
